"""
Main test runner.
Executes all tests and generates Excel report.
"""

import unittest
import sys
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from .system.system_tests import TestChessGameSystem
from .integration.integration_tests import TestChessGameIntegration
from .unit.unit_tests import TestChessGameUnit


class ChessTestResult(unittest.TestResult):
    """
    Custom test result class to capture detailed information for Excel export.
    Tracks each test case with its ID, description, result, and execution details.
    """
    def __init__(self):
        super().__init__()
        self.test_details = []
        
    def startTest(self, test):
        """Called when a test starts - capture start time."""
        super().startTest(test)
        self.current_start_time = datetime.now()
        
    def addSuccess(self, test):
        """Called when a test passes."""
        super().addSuccess(test)
        self._add_test_detail(test, 'PASS', '')
        
    def addFailure(self, test, err):
        """Called when a test fails (assertion error)."""
        super().addFailure(test, err)
        error_msg = self._format_error(err)
        self._add_test_detail(test, 'FAIL', error_msg)
        
    def addError(self, test, err):
        """Called when a test has an error (exception)."""
        super().addError(test, err)
        error_msg = self._format_error(err)
        self._add_test_detail(test, 'ERROR', error_msg)
        
    def addSkip(self, test, reason):
        """Called when a test is skipped."""
        super().addSkip(test, reason)
        self._add_test_detail(test, 'SKIP', reason)
        
    def _add_test_detail(self, test, result, error_msg):
        """
        Extract test details and add to results list.
        Looks for test_id and test_description attributes on test methods.
        """
        end_time = datetime.now()
        duration = (end_time - self.current_start_time).total_seconds()
        
        # Get test case ID and description from test method attributes
        test_method = getattr(test, test._testMethodName)
        test_id = getattr(test_method, 'test_id', 'TC-UNKNOWN')
        description = getattr(test_method, 'test_description', test._testMethodName)
        
        # Determine test category from test class name
        class_name = test.__class__.__name__
        if 'System' in class_name:
            category = 'SYSTEM'
        elif 'Integration' in class_name:
            category = 'INTEGRATION'
        elif 'Unit' in class_name:
            category = 'UNIT'
        else:
            category = 'OTHER'
        
        self.test_details.append({
            'test_id': test_id,
            'category': category,
            'test_name': test._testMethodName,
            'description': description,
            'result': result,
            'error_message': error_msg,
            'duration': duration,
            'timestamp': end_time.strftime('%Y-%m-%d %H:%M:%S')
        })
        
    def _format_error(self, err):
        """Format error tuple into readable string."""
        import traceback
        return ''.join(traceback.format_exception(*err))


def generate_excel_report(test_result, output_path):
    """
    Generate Excel report from test results.
    
    Args:
        test_result: ChessTestResult object with test details
        output_path: Path where Excel file should be saved
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Results"
    
    # Define styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    error_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    skip_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Define headers
    headers = ['Test Case ID', 'Category', 'Test Name', 'Description', 
               'Result', 'Duration (s)', 'Timestamp', 'Error Message']
    
    # Set column widths
    ws.column_dimensions['A'].width = 15  # Test Case ID
    ws.column_dimensions['B'].width = 12  # Category
    ws.column_dimensions['C'].width = 30  # Test Name
    ws.column_dimensions['D'].width = 50  # Description
    ws.column_dimensions['E'].width = 10  # Result
    ws.column_dimensions['F'].width = 12  # Duration
    ws.column_dimensions['G'].width = 20  # Timestamp
    ws.column_dimensions['H'].width = 60  # Error Message
    
    # Write headers
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    # Write test data
    for idx, test in enumerate(test_result.test_details, start=2):
        # Test Case ID
        cell = ws.cell(row=idx, column=1, value=test['test_id'])
        cell.alignment = center_align
        cell.border = thin_border
        
        # Category
        cell = ws.cell(row=idx, column=2, value=test['category'])
        cell.alignment = center_align
        cell.border = thin_border
        
        # Test Name
        cell = ws.cell(row=idx, column=3, value=test['test_name'])
        cell.alignment = left_align
        cell.border = thin_border
        
        # Description
        cell = ws.cell(row=idx, column=4, value=test['description'])
        cell.alignment = left_align
        cell.border = thin_border
        
        # Result with color coding
        result = test['result']
        cell = ws.cell(row=idx, column=5, value=result)
        cell.alignment = center_align
        cell.border = thin_border
        if result == 'PASS':
            cell.fill = pass_fill
        elif result == 'FAIL':
            cell.fill = fail_fill
        elif result == 'ERROR':
            cell.fill = error_fill
        elif result == 'SKIP':
            cell.fill = skip_fill
        
        # Duration
        cell = ws.cell(row=idx, column=6, value=f"{test['duration']:.3f}")
        cell.alignment = center_align
        cell.border = thin_border
        
        # Timestamp
        cell = ws.cell(row=idx, column=7, value=test['timestamp'])
        cell.alignment = center_align
        cell.border = thin_border
        
        # Error Message
        cell = ws.cell(row=idx, column=8, value=test['error_message'])
        cell.alignment = left_align
        cell.border = thin_border
        ws.row_dimensions[idx].height = 60 if test['error_message'] else 20
    
    # Add summary sheet
    summary_ws = wb.create_sheet(title="Summary")
    
    # Calculate statistics
    total = len(test_result.test_details)
    passed = sum(1 for t in test_result.test_details if t['result'] == 'PASS')
    failed = sum(1 for t in test_result.test_details if t['result'] == 'FAIL')
    errors = sum(1 for t in test_result.test_details if t['result'] == 'ERROR')
    skipped = sum(1 for t in test_result.test_details if t['result'] == 'SKIP')
    pass_rate = (passed / total * 100) if total > 0 else 0
    
    # Write summary
    summary_data = [
        ['Test Execution Summary', ''],
        ['', ''],
        ['Total Tests', total],
        ['Passed', passed],
        ['Failed', failed],
        ['Errors', errors],
        ['Skipped', skipped],
        ['Pass Rate', f'{pass_rate:.2f}%'],
        ['', ''],
        ['Execution Date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
    ]
    
    for row_idx, (label, value) in enumerate(summary_data, start=1):
        summary_ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        summary_ws.cell(row=row_idx, column=2, value=value)
    
    summary_ws.column_dimensions['A'].width = 20
    summary_ws.column_dimensions['B'].width = 20
    
    # Save workbook
    wb.save(output_path)
    print(f"\nExcel report saved to: {output_path}")


def run_tests():
    """
    Main function to run all tests and generate Excel report.
    """
    # Create test suite
    loader = unittest.TestLoader()
    suite = unittest.TestSuite()
    
    # Add all test classes
    suite.addTests(loader.loadTestsFromTestCase(TestChessGameSystem))
    suite.addTests(loader.loadTestsFromTestCase(TestChessGameIntegration))
    suite.addTests(loader.loadTestsFromTestCase(TestChessGameUnit))
    
    # Run tests with custom result class
    print("Running Chess Game Tests...\n")
    result = ChessTestResult()
    suite.run(result)
    
    # Print summary to console
    print("\n" + "="*70)
    print("TEST EXECUTION SUMMARY")
    print("="*70)
    print(f"Total Tests: {result.testsRun}")
    print(f"Passed: {len([t for t in result.test_details if t['result'] == 'PASS'])}")
    print(f"Failed: {len(result.failures)}")
    print(f"Errors: {len(result.errors)}")
    print(f"Skipped: {len(result.skipped)}")
    print("="*70)
    
    # Generate Excel report
    export_dir = os.path.join(os.path.dirname(__file__), 'export')
    os.makedirs(export_dir, exist_ok=True)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_path = os.path.join(export_dir, f'chess_test_results_{timestamp}.xlsx')
    
    generate_excel_report(result, output_path)
    
    return result


if __name__ == '__main__':
    result = run_tests()
    # Exit with error code if tests failed
    sys.exit(0 if result.wasSuccessful() else 1)